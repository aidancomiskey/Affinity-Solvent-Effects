from titration_class import TitrationGroup
from tqdm import tqdm
from openpyxl import load_workbook

# Performs curve fitting and plotting for all titrations of exp_number listed in titration_list
def analyze_all_titrations(titration_list, experimental_params_file, fit_params_file, experimental_values_file,
                           spectrum_file_format, correct_baseline=False, print_report=False,
                           baseline_wavelength=750, baseline_absorbance=0):
    titration_group = TitrationGroup(experimental_params_file=experimental_params_file,
                                     fit_params_file=fit_params_file,
                                     experimental_values_file=experimental_values_file,
                                     spectrum_file_format=spectrum_file_format)
    for titration in tqdm(titration_list):
        titration_group.load_titration(titration)
        if correct_baseline:
            titration_group[titration].correct_baselines(ref_wavelength=baseline_wavelength,
                                                         ref_absorbance=baseline_absorbance)
        titration_group[titration].analyze_titration(print_report=print_report)
    titration_group.get_affinities()
    return titration_group


# Use after updating selected wavelengths, before initializing TitrationGroup
def update_direct_exp_values_excel(direct_titrations, experimental_params_file, fit_params_file,
                                   experimental_values_file, spectrum_file_format, correct_baseline, print_report):
    update_group = analyze_all_titrations(titration_list=direct_titrations,
                                          experimental_params_file=experimental_params_file,
                                          fit_params_file=fit_params_file,
                                          experimental_values_file=experimental_values_file,
                                          spectrum_file_format=spectrum_file_format,
                                          correct_baseline=correct_baseline, print_report=print_report)
    workbook = load_workbook(filename=update_group.experimental_values_filename)
    sheet = workbook["Values"]
    n_rows = sheet.max_row
    if n_rows != len(direct_titrations) + 1:
        raise ValueError("Direct titrations list does not match those in excel file")
    for i in range(2, n_rows + 1):
        experiment_number = str(sheet[f'A{i}'].value)
        sheet[f'B{i}'] = update_group[experiment_number].fit_output.params["epsilon_host1"].value
        sheet[f'C{i}'] = update_group[experiment_number].fit_output.params["epsilon_complex1"].value
        sheet[f'D{i}'] = update_group[experiment_number].fit_output.params["epsilon_host2"].value
        sheet[f'E{i}'] = update_group[experiment_number].fit_output.params["epsilon_complex2"].value
        sheet[f'F{i}'] = update_group[experiment_number].fit_output.params["k_i"].value
    workbook.save(filename=update_group.experimental_values_filename)
    print(f'Coefficients updated!       See {update_group.experimental_values_filename}')
